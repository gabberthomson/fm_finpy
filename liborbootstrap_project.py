from datetime import date
from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('LiborCurveBootstrap.xlsx')
    ws = wb.active

    # Take the input parameters
    today = ws['C2'].value.date()
    libor_tenor = ws['C3'].value
    libor_value = ws['C4'].value

    # OIS Data
    ois_startdate = today
    ois_maturities = []
    ois_mktquotes = []
    for cell in list(ws.iter_rows('B8:C37')):
        ois_maturities.append(cell[0].value)
        ois_mktquotes.append(cell[1].value)

    # Swap Data
    swap_startdate = today
    swap_maturities = []
    swap_mktquotes = []
    for cell in list(ws.iter_rows('E8:F40')):
        swap_maturities.append(cell[0].value)
        swap_mktquotes.append(cell[1].value)

    # Output Dates
    output_dates = []
    for cell in list(ws.iter_rows('H8:H40')):
        output_dates.append(cell[0].value.date())


    # YOUR CODE HERE .... The result of your code must be a variable of type list whose name
    # must be output_results. The length of this list has to be the same of output_dates




    # END OF YOUR CODE

    # Write results
    # A variable named output_results of type list, with the same length of output_dates, is expected.
    # In case this is not present, a message is written
    if 'output_results' not in locals():
        output_result = ["Not Successful" for x in range(len(output_dates))]

    out_list = list(ws.iter_rows('I8:I40'))
    for i in range(len(output_result)):
        out_list[i][0].value = output_result[i]

    # A new file with the results is created
    wb.save("LiborCurveBootstrap_output.xlsx")