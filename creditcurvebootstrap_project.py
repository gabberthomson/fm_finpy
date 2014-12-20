from datetime import date
from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('CreditCurveBootstrap.xlsx')
    ws = wb.active

    # Take the input parameters
    today = ws['C2'].value.date()
    cds_recovery = ws['C3'].value
    cds_startdate = today # We assume that all of the cds start today

    # OIS Data
    ois_startdate = today
    ois_maturities = []
    ois_mktquotes = []
    for cell in list(ws.iter_rows('B16:C45')):
        ois_maturities.append(cell[0].value)
        ois_mktquotes.append(cell[1].value)

    # CDS Input Data
    cds_input_maturities = []
    cds_input_quotes = []
    for cell in list(ws.iter_rows('B7:C12')):
        cds_input_maturities.append(cell[0].value)
        cds_input_quotes.append(cell[1].value)

    # CDS Output Data
    nominals = []
    cds_types = []
    maturities = []
    spreads = []

    for cell in list(ws.iter_rows('E5:H19')):
        nominals.append(cell[0].value)
        cds_types.append(cell[1].value)
        maturities.append(cell[2].value)
        spreads.append(cell[3].value)


    # YOUR CODE HERE ....
    # In the calculation of the npv of output CDS, please note that you have to take into account
    # both the nominal and the type, i.e if the present value is seen from the protection seller or
    # the protection buyer
    # The result of your code must be a variables of type list named
    # output_npv. The length of this list has to be the equal to the number of output cds
    # i.e len(nominals) for example




    # END OF YOUR CODE

    # Write results
    # A variable named output_results of type list, with the same length of output_dates, is expected.
    # In case this is not present, a message is written
    if 'output_npv' not in locals():
        output_npv = ["Not Successful" for x in range(len(nominals))]

    out_list = list(ws.iter_rows('I5:I19'))
    for i in range(len(output_npv)):
        out_list[i][0].value = output_npv[i]

    # A new file with the results is created
    wb.save("CreditCurveBootstrap_output.xlsx")