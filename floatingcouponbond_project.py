from datetime import date
from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('FloatingCouponBond.xlsx')
    ws = wb.active

    # Take the input parameters
    today = ws['C2'].value.date()

    libor_tenor = ws['C3'].value
    # Please note that the payment frequency is the same of the libor tenor

    # OIS Data
    ois_startdate = today
    ois_maturities = []
    ois_mktquotes = []
    for cell in list(ws.iter_rows('B17:C46')):
        ois_maturities.append(cell[0].value)
        ois_mktquotes.append(cell[1].value)

    # Credit Curve Data
    ndps = []
    ndpdates = []
    for cell in list(ws.iter_rows('B8:C13')):
        ndpdates.append(cell[0].value.date())
        ndps.append(cell[1].value)

    # Bond data
    nominals = []
    start_dates = []
    end_dates = []
    current_coupon = []
    margin = []
    recovery_rates = []

    for cell in list(ws.iter_rows('E5:J19')):
        nominals.append(cell[0].value)
        start_dates.append(cell[1].value.date())
        end_dates.append(cell[2].value.date())
        current_coupon.append(cell[3].value)
        margin.append(cell[4].value)
        recovery_rates.append(cell[5].value)


    # YOUR CODE HERE ....
    # In the coupon calculation use act/360 convention to compute the accrual period (i.e. tau)
    # The result of your code must be a variables of type list named
    # output_npv. The length of this list has to be the equal to the number of bonds
    # i.e len(nominals) for example




    # END OF YOUR CODE

    # Write results
    # A variable named output_results of type list, with the same length of output_dates, is expected.
    # In case this is not present, a message is written
    if 'output_npv' not in locals():
        output_npv = ["Not Successful" for x in range(len(nominals))]

    out_list = list(ws.iter_rows('K5:K19'))
    for i in range(len(output_npv)):
        out_list[i][0].value = output_npv[i]

    # A new file with the results is created
    wb.save("FloatingCouponBond_output.xlsx")