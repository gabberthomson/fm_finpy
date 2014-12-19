from datetime import date
from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('FixedCouponBond.xlsx')
    ws = wb.active

    # Take the input parameters
    today = ws['C2'].value.date()

    # OIS Data
    ois_startdate = today
    ois_maturities = []
    ois_mktquotes = []
    for cell in list(ws.iter_rows('B16:C45')):
        ois_maturities.append(cell[0].value)
        ois_mktquotes.append(cell[1].value)

    # Credit Curve Data
    ndps = []
    ndpdates = []
    for cell in list(ws.iter_rows('B6:C11')):
        ndpdates.append(cell[0].value)
        ndps.append(cell[1].value)

    # Bond data
    nominals = []
    start_dates = []
    end_dates = []
    cpn_frequency = []
    coupons = []
    recovery_rates = []

    for cell in list(ws.iter_rows('E5:J19')):
        nominals.append(cell[0].value)
        start_dates.append(cell[1].value)
        end_dates.append(cell[2].value)
        cpn_frequency.append(cell[3].value)
        coupons.append(cell[4].value)
        recovery_rates.append(cell[5].value)


    # YOUR CODE HERE .... The result of your code must be a variables of type list named
    # output_npv. The length of this list has to be the equal to the number of bonds
    # i.e len(nominals) for example




    # END OF YOUR CODE

    # Write results
    # A variable named output_results of type list, with the same length of output_dates, is expected.
    # In case this is not present, a message is written
    if 'output_npv' not in locals():
        output_npv = ["Not Successful" for x in range(len(nominals))]

    out_list = list(ws.iter_rows('K5:K19'))
    for i in range(len(output_npv)):
        out_list[i][0].value = output_npv[i]

    # A new file with the results is created
    wb.save("FixedCouponBond_output.xlsx")